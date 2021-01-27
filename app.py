import openpyxl

# wb = openpyxl.Workbook()
# this creates an empty woorkbook with one sheet in memory.
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

# wb.create_sheet("Sheet2", 0)
# wb.remove_sheet(sheet)

cell = sheet["a1"]
column = sheet["a"]
cells = sheet["a:c"]

sheet.append([1004, 4, 10])
wb.save("transactions2.xlsx")
